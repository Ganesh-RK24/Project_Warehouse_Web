#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import io
import csv
import datetime
from flask import Flask, request, redirect, url_for, session, render_template_string, flash
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from dateutil import parser as date_parser
import secrets


# In[2]:


app = Flask(__name__)
app.secret_key = secrets.token_hex(24)
app.permanent_session_lifetime = datetime.timedelta(minutes=30)

ALLOWED_EXTENSIONS = {'csv'}

USERS = {
    "vendikatla.s@matrixbsindia.com": {
        "password": "matrix@123",
        "folder": "Matrix Business Services India Private Limited",
        "name": "Subhan"
    },
    "deekshith@proteam.co.in": {
        "password": "proteam@123",
        "folder": "PRO-Team Solution Pvt Ltd",
        "name": "Deekshith"
    }
}

ONEDRIVE_BASE = r"D:\OneDrive - Aditya Birla Fashion & Retail Ltd\FY_2025-26\Project_Warehouse Audit\AuditorBins"
LOG_XLSX_PATH = r"D:\OneDrive - Aditya Birla Fashion & Retail Ltd\FY_2025-26\upload_logs.xlsx"


# In[3]:


LOGIN_HTML = """
<!doctype html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Login - Auditor Portal</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Roboto&display=swap');
  body {
    margin: 0; font-family: 'Roboto', sans-serif;
    background: #121212;
    color: #eee;
    display: flex;
    justify-content: center;
    align-items: center;
    height: 100vh;
  }
  .login-container {
    background: #1f1f1f;
    padding: 40px 30px 50px 30px;
    border-radius: 12px;
    width: 320px;
    box-shadow: 0 0 15px #4a90e2;
    text-align: center;
  }
  h2 {
    margin-bottom: 30px;
    font-weight: 700;
    color: #4a90e2;
  }
  label {
    display: block;
    margin-bottom: 8px;
    font-weight: 500;
    text-align: left;
  }
  input[type="email"],
  input[type="password"] {
    width: 100%;
    padding: 10px 12px;
    margin-bottom: 20px;
    border: none;
    border-radius: 6px;
    outline: none;
    font-size: 1rem;
    background: #333;
    color: #eee;
  }
  input[type="email"]:focus,
  input[type="password"]:focus {
    box-shadow: 0 0 8px 2px #4a90e2;
  }
  input[type="submit"] {
    width: 100%;
    padding: 12px;
    background-color: #4a90e2;
    border: none;
    border-radius: 6px;
    font-size: 1rem;
    font-weight: 700;
    color: #fff;
    cursor: pointer;
    transition: background-color 0.3s ease;
  }
  input[type="submit"]:hover {
    background-color: #357ABD;
  }
  ul {
    padding-left: 20px;
    margin-top: 15px;
  }
  ul li {
    margin-bottom: 8px;
    color: #ff6b6b;
  }
</style>
</head>
<body>

<div class="login-container">
  <h2>Auditor Login</h2>
  <form method="POST" novalidate>
    <label for="email">Email</label>
    <input type="email" name="email" id="email" placeholder="you@example.com" required>

    <label for="password">Password</label>
    <input type="password" name="password" id="password" placeholder="Enter your password" required>

    <input type="submit" value="Login">
  </form>
  {% with messages = get_flashed_messages() %}
    {% if messages %}
      <ul>
        {% for msg in messages %}
          <li>{{ msg }}</li>
        {% endfor %}
      </ul>
    {% endif %}
  {% endwith %}
</div>

</body>
</html>
"""


# In[4]:


UPLOAD_HTML = """
<!doctype html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Upload CSV - Auditor Portal</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Roboto&display=swap');
  body {
    margin: 0; font-family: 'Roboto', sans-serif;
    background: #121212;
    color: #eee;
    min-height: 100vh;
    display: flex;
    justify-content: center;
    align-items: center;
  }
  .upload-container {
    background: #1f1f1f;
    padding: 35px 40px 50px 40px;
    border-radius: 12px;
    width: 550px;
    box-shadow: 0 0 25px #4a90e2;
    text-align: center;
  }
  h2 {
    margin-top: 0;
    margin-bottom: 25px;
    font-weight: 700;
    color: #4a90e2;
  }
  form {
    display: flex;
    flex-direction: column;
    align-items: center;
  }
  input[type="file"], input[type="submit"], input[type="number"], input[type="date"], select {
    margin-bottom: 20px;
    padding: 10px;
    border-radius: 6px;
    border: none;
    background-color: #333;
    color: #eee;
    width: 100%;
    box-sizing: border-box;
  }
  input[type="file"]::-webkit-file-upload-button {
    cursor: pointer;
    background: #4a90e2;
    border: none;
    padding: 8px 14px;
    border-radius: 6px;
    color: #fff;
    font-weight: 700;
  }
  input[type="submit"] {
    background-color: #4a90e2;
    color: white;
    font-weight: 700;
    border: none;
    cursor: pointer;
  }
  input[type="submit"]:hover,
  input[type="file"]::-webkit-file-upload-button:hover {
    background-color: #357ABD;
  }
  p.logout-link {
    margin-top: 20px;
  }
  p.logout-link a {
    color: #4a90e2;
    text-decoration: underline;
    font-weight: 600;
  }
  .date-row {
    display: flex;
    gap: 10px;
    margin-bottom: 15px;
    width: 100%;
  }
  .date-row input, .date-row select {
    flex: 1;
  }
  .add-button {
    background-color: transparent;
    color: #4a90e2;
    border: 2px dashed #4a90e2;
    font-size: 1.5rem;
    cursor: pointer;
    padding: 6px 10px;
    border-radius: 6px;
    width: 100%;
  }
  ul {
    margin-top: 15px;
    padding-left: 18px;
    text-align: left;
  }
  ul.success {
    color: #4caf50;
  }
  ul.error {
    color: #ff6b6b;
  }
</style>
<script>
  function addDateRow() {
    const container = document.getElementById('date-rows');
    const row = document.createElement('div');
    row.className = 'date-row';

    row.innerHTML = `
      <input type="date" name="entry_date[]" required>
      <select name="warehouse_no[]" required>
        <option value="">Select WH No</option>
        <option value="W501">W501</option>
        <option value="W503">W503</option>
        <option value="W504">W504</option>
        <option value="W506">W506</option>
        <option value="W509">W509</option>
        <option value="W511">W511</option>
        <option value="W512">W512</option>
      </select>
      <input type="number" name="manpower[]" placeholder="Manpower Count (optional)" min="0">
    `;
    container.appendChild(row);
  }
</script>
</head>
<body>

<div class="upload-container">
  <h2>Upload CSV - Logged in as {{ user_name }}</h2>
  <form method="post" enctype="multipart/form-data" novalidate>
    <input type="file" name="file" accept=".csv" required>

    <div id="date-rows">
      <div class="date-row">
        <input type="date" name="entry_date[]" required>
        <select name="warehouse_no[]" required>
          <option value="">Select WH No</option>
          <option value="W501">W501</option>
          <option value="W503">W503</option>
          <option value="W504">W504</option>
          <option value="W506">W506</option>
          <option value="W509">W509</option>
          <option value="W511">W511</option>
          <option value="W512">W512</option>
        </select>
        <input type="number" name="manpower[]" placeholder="Manpower Count (optional)" min="0">
      </div>
    </div>

    <button type="button" class="add-button" onclick="addDateRow()">+ Add More Dates</button>

    <input type="submit" value="Upload">
  </form>
  <p class="logout-link"><a href="/logout">Logout</a></p>

  {% with messages = get_flashed_messages() %}
    {% if messages %}
      <ul class="{% if '❌' in messages[0] %}error{% else %}success{% endif %}">
        {% for msg in messages %}
          <li>{{ msg }}</li>
        {% endfor %}
      </ul>
    {% endif %}
  {% endwith %}
</div>

</body>
</html>
"""


# In[5]:


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_csv_by_wh(file_stream, vendor_path):
    file_stream.seek(0)
    decoded_stream = io.StringIO(file_stream.read().decode('utf-8'))
    reader = csv.reader(decoded_stream)

    raw_headers = next(reader)
    headers = [h.strip() for h in raw_headers]

    wh_col_index = None
    date_col_index = None

    for idx, h in enumerate(headers):
        col = h.strip().lower()
        if col == "warehouse no":
            wh_col_index = idx
        if col == "counted date":
            date_col_index = idx

    if wh_col_index is None:
        raise ValueError("❌ 'Warehouse No' column not found in uploaded file.")
    if date_col_index is None:
        raise ValueError("❌ 'Counted Date' column not found in uploaded file.")

    wh_groups = {}
    wh_dates = {}

    for row in reader:
        if len(row) <= max(wh_col_index, date_col_index):
            continue
        wh = row[wh_col_index].strip()
        date_raw = row[date_col_index].strip()

        if not wh or not date_raw:
            continue

        try:
            counted_date = date_parser.parse(date_raw, dayfirst=True).date()
        except Exception:
            continue

        wh_groups.setdefault(wh, []).append(row)
        wh_dates.setdefault(wh, []).append(counted_date)

    if not wh_groups:
        raise ValueError("❌ No valid data grouped by warehouse.")

    saved_files = []

    for wh, rows in wh_groups.items():
        wh_folder = os.path.join(vendor_path, wh)
        os.makedirs(wh_folder, exist_ok=True)

        dates = wh_dates.get(wh, [])
        if not dates:
            continue
        low_date = min(dates).strftime('%Y-%m-%d')
        high_date = max(dates).strftime('%Y-%m-%d')

        filename = f"WH_Counted_{wh}_{low_date}-{high_date}.csv"
        save_path = os.path.join(wh_folder, filename)

        with open(save_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

        saved_files.append((wh, save_path))

    return saved_files


# In[6]:


def log_upload(user_email, filename, warehouse_list):
    now = datetime.datetime.now()
    wb = None
    if os.path.exists(LOG_XLSX_PATH):
        wb = load_workbook(LOG_XLSX_PATH)
    else:
        wb = Workbook()

    ws = wb.active
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(["Date", "Time", "User Email", "Filename", "Warehouses"])

    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        user_email,
        filename,
        ", ".join(warehouse_list)
    ])

    wb.save(LOG_XLSX_PATH)


# In[7]:


@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user' in session:
        return redirect(url_for('upload_file'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        if email in USERS and USERS[email]['password'] == password:
            session['user'] = email
            session.permanent = True
            return redirect(url_for('upload_file'))
        else:
            flash("Invalid email or password.")

    return render_template_string(LOGIN_HTML)

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if 'user' not in session:
        return redirect(url_for('login'))

    email = session['user']
    user_info = USERS.get(email)
    if not user_info:
        flash("Unauthorized.")
        return redirect(url_for('logout'))

    vendor_path = os.path.join(ONEDRIVE_BASE, user_info['folder'])
    os.makedirs(vendor_path, exist_ok=True)

    if request.method == 'POST':
        if 'file' not in request.files or request.files['file'].filename == '':
            flash("❌ File upload is mandatory.")
            return redirect(request.url)

        file = request.files['file']

        dates = request.form.getlist('entry_date[]')
        warehouse_nos = request.form.getlist('warehouse_no[]')
        manpower_counts = request.form.getlist('manpower[]')

        if not dates or not warehouse_nos or len(dates) != len(warehouse_nos):
            flash("❌ Please fill all Date and Warehouse No fields.")
            return redirect(request.url)

        for d, wh in zip(dates, warehouse_nos):
            if d.strip() == '' or wh.strip() == '':
                flash("❌ Date and Warehouse No fields cannot be empty.")
                return redirect(request.url)

        for i in range(len(manpower_counts), len(dates)):
            manpower_counts.append('')

        manpower_data = list(zip(dates, warehouse_nos, manpower_counts))

        if file and allowed_file(file.filename):
            try:
                saved_files = save_csv_by_wh(file.stream, vendor_path)
                wh_list = [wh for wh, _ in saved_files]

                # Log upload with WH numbers
                log_upload(email, file.filename, wh_list)

                # Save manpower details CSV in vendor folder
                manpower_csv_path = os.path.join(vendor_path, "Manpower Details.csv")
                file_exists = os.path.exists(manpower_csv_path)

                with open(manpower_csv_path, 'a', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    if not file_exists:
                        writer.writerow(["Date", "Warehouse No", "Manpower Count"])
                    for date_val, wh_no, manpower in manpower_data:
                        writer.writerow([date_val, wh_no, manpower.strip() if manpower else ''])

                flash(f"✅ File uploaded and saved under: {', '.join(wh_list)}")
                flash(f"✅ Manpower details saved.")

            except Exception as e:
                flash(f"❌ Error processing file: {e}")
        else:
            flash("❌ Only CSV files are allowed.")

    return render_template_string(UPLOAD_HTML, user_name=user_info['name'], email=email)


# In[ ]:


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

